from __future__ import print_function
import time
import requests
import json
import intersight
from intersight.intersight_api_client import IntersightApiClient
from openpyxl import Workbook

# Creates Intersight API instance
api_instance = IntersightApiClient(
    private_key="./keyfile.txt",
    api_key_id=open('./apikey.txt',"r").read(),
)

# Initiates the Intersight API connectors
ComputeRackUnitApi = intersight.ComputeRackUnitApi(api_instance)
EquipmentPsuApi = intersight.EquipmentPsuApi(api_instance)
BiosUnitApi	= intersight.BiosUnitApi(api_instance)
FirmwareRunningFirmwareApi = intersight.FirmwareRunningFirmwareApi(api_instance)
ManagementControllerApi = intersight.ManagementControllerApi(api_instance)
ComputeBoardApi = intersight.ComputeBoardApi(api_instance)
ProcessorUnitApi = intersight.ProcessorUnitApi(api_instance)
MemoryArrayApi = intersight.MemoryArrayApi(api_instance)
MemoryUnitApi = intersight.MemoryUnitApi(api_instance)
PciDeviceApi = intersight.PciDeviceApi(api_instance)
AdapterUnitApi = intersight.AdapterUnitApi(api_instance)
StorageControllerApi = intersight.StorageControllerApi(api_instance)
StoragePhysicalDiskApi = intersight.StoragePhysicalDiskApi(api_instance)
    
wb = Workbook()

# Creates the worksheets
ws_server = wb.active
ws_server.title = "Server"
ws_memory = wb.create_sheet("Memory")
#ws_networkadapter = wb.create_sheet("Network Adapter")
ws_pcie = wb.create_sheet("PCIe Devices")
ws_disk = wb.create_sheet("Physical Disks")
ws_psu = wb.create_sheet("PSUs")

# Creates table headers
ws_server.append(['Server ID', 'User Label', 'Asset Tag', 'Serial', 'DN', 'Model', 'Oper Power State', 'Operability', 'MOID', 'UUID', 'Platform Type', 'Service Profile', 'KVM IP', 'Num of CPUs', 'CPU Model', 'CPU Architecture', 'Num of CPU Cores', 'Num of CPU Cores enabled', 'Num of Threads', 'CPU Speed', 'CPU Stepping', 'Total Memory', 'Memory Speed', 'Max Memory Units', 'Num of Adaptors', 'Num of Eth Host Int', 'Num of FC Host Int', 'Compute Board Serial', 'BIOS Firmware Version', 'BMC Firmware System', 'BMC Firmware Bootloader'])
ws_memory.append(['Serial', 'DN', 'Server MOID (Parent)', 'Presence', 'Operational State', 'Model', 'MOID',  'Form Factor', 'Type', 'Capacity', 'Clock', 'Speed', 'Latency', 'Bank', 'Location', 'Memory ID'])
#ws_networkadapter.append([])
ws_pcie.append(['MOID', 'DN', 'Server MOID (Parent)', 'Model', 'Slot ID', 'Firmware Version'])
ws_disk.append(['Serial', 'DN', 'Server MOID (Parent)', 'MOID', 'Disk ID', 'Type', 'Vendor', 'Model', 'Protocol', 'Size', 'Link Speed'])
ws_psu.append(['Serial', 'DN', 'Server MOID (Parent)', 'Model', 'MOID', 'PSU ID', 'PSU Presence'])

racks = ComputeRackUnitApi.compute_rack_units_get().results

rack_count = len(racks)
current_rack = 0
for r in racks:
    state = 'Processing Rack Servers: ' + str(current_rack) + '/' + str(rack_count) + ' done'
    print (state, end="\r")
    
    bios_fw = FirmwareRunningFirmwareApi.firmware_running_firmwares_moid_get(BiosUnitApi.bios_units_moid_get(r.biosunits[0].moid).running_firmware[0].moid)
    bmc = ManagementControllerApi.management_controllers_moid_get(r.bmc.moid).running_firmware
    
    bmc_fw = {
        'system': None,
        'boot-loader': None
    }
    for b in bmc:
        x = FirmwareRunningFirmwareApi.firmware_running_firmwares_moid_get(b.moid)
        if(x.component == 'boot-loader'):
            bmc_fw['boot-loader'] = x.version
        elif(x.component == 'system'):
            bmc_fw['system'] = x.version
            
    compute_board = ComputeBoardApi.compute_boards_moid_get(r.board.moid)
    p = ProcessorUnitApi.processor_units_moid_get(compute_board.processors[0].moid)
    
    memory = MemoryArrayApi.memory_arrays_moid_get(compute_board.memory_arrays[0].moid)
    for u in memory.units:
        mem_unit = MemoryUnitApi.memory_units_moid_get(u.moid)
        ws_memory.append([mem_unit.serial, mem_unit.dn, r.moid, mem_unit.presence, mem_unit.oper_state, mem_unit.operability, mem_unit.model, mem_unit.moid, mem_unit.form_factor, mem_unit.type, mem_unit.capacity, mem_unit.clock, mem_unit.speed, mem_unit.latency, mem_unit.bank, mem_unit.location, mem_unit.memory_id])
        
    #print(compute_board.presence)
    
    ws_server.append([r.server_id, r.user_label, r.asset_tag, r.serial, r.dn, r.model, r.oper_power_state, r.moid, r.uuid, r.platform_type, r.service_profile, r.kvm_ip_addresses[0].address, r.num_cpus, p.model, p.architecture, r.num_cpu_cores, r.num_cpu_cores_enabled, r.num_threads, p.speed, p.stepping, r.total_memory, r.memory_speed, memory.max_devices, r.num_adaptors, r.num_eth_host_interfaces, r.num_fc_host_interfaces, compute_board.serial, bios_fw.version, bmc_fw['system'], bmc_fw['boot-loader']])
    
    #print(r.mod_time)
    #print(r.board)
    #print(r.tags)
    #print(r.locator_led)
    #print(r.registered_device)
    #print(r.rack_enclosure_slot)
    #print(r.storage_enclosures)
    #print(r.object_type)
    #print(r.fanmodules)
    #print(r.sas_expanders)
    #print(r.device_mo_id)
    #print(r.generic_inventory_holders)
    #print(bios_fw.package_version)
    #print(bios_fw.registered_device)
    #print(compute_board.graphics_cards)
    #print(compute_board.equipment_tpms)
    #print(compute_board.storage_flex_flash_controllers)
    #print(compute_board.storage_flex_util_controllers)
    #for a in r.adapters:
        #print(AdapterUnitApi.adapter_units_moid_get(a.moid))
    
    for p in r.psus:
        psu = EquipmentPsuApi.equipment_psus_moid_get(p.moid)
        ws_psu.append([psu.serial, psu.dn, r.moid, psu.model, psu.moid, psu.psu_id, psu.presence])
        
        #print(psu.equipment_chassis)
        #print(psu.equipment_rack_enclosure)
        #print(psu.mod_time)
        #print(psu.device_mo_id)
    
    for d in r.pci_devices:
        pcid = PciDeviceApi.pci_devices_moid_get(d.moid)
        ws_pcie.append([pcid.moid, pcid.dn, r.moid, pcid.model, pcid.slot_id, pcid.firmware_version])
        
        #print(pcid.serial)
        #print(pcid.vendor)
        
    for s in compute_board.storage_controllers:
        sc = StorageControllerApi.storage_controllers_moid_get(s.moid)
#        print(sc.controller_id)
#        print(sc.controller_status)
#        print(sc.dn)
#        print(sc.model)
#        print(sc.moid)
#        print(sc.pci_addr)
#        print(sc.pci_slot)
        for d in sc.physical_disks:
            pdisk = StoragePhysicalDiskApi.storage_physical_disks_moid_get(d.moid)
            ws_disk.append([pdisk.serial, pdisk.dn, r.moid, pdisk.disk_id, pdisk.moid, pdisk.type, pdisk.vendor, pdisk.model, pdisk.protocol, pdisk.size, pdisk.link_speed])
            
            #print(pdisk.block_size)
            #print(pdisk.bootable)
            #print(pdisk.disk_state)
            #print(pdisk.link_state)
            #print(pdisk.physical_block_size)
            #print(pdisk.presence)
            #print(pdisk.raw_size)
            #print(pdisk.running_firmware)
               
#        print(sc.presence)
#        print(sc.raid_support)
#        print(sc.rebuild_rate)
#        print(sc.running_firmware)
#        print(sc.serial)
#        print(sc.type)
#        print(sc.vendor)
#        print(sc.virtual_drives)

    current_rack += 1
    
    
# Save the Excel file
wb.save("intersight_report.xlsx")